import openpyxl
import pytest
from selenium.webdriver.common.by import By
from selenium.webdriver.support.expected_conditions import visibility_of_element_located
import logging
logger = logging.getLogger(__name__)
class TestCase1:

    print("testcase1")
    @pytest.fixture(autouse=True, scope="session")
    def test_case_login(self, browser):
        # 加载Excel文件
        workbook = openpyxl.load_workbook('test_excel.xlsx')
        sheet = workbook.active
        # 遍历Excel文件中的测试用例
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # 读取测试用例信息
            case_id = int(row[0])
            action = row[3]
            param1 = row[4]
            param2 = row[5]
            if case_id == 1:
                # 根据关键字执行相应的操作
                if action == 'name':
                    print(f"Test Case {case_id}: {param1}")
                elif action == 'goto':
                    browser.goto(param1)
                elif action == 'input':
                    browser.input(param1, param2)
                elif action == 'click':
                    browser.click(param1)
                elif action == 'hide_div':
                    browser.hide_div(param1, param2)
            else:
                continue
        browser.driver.get_screenshot_as_file("case_login.png")
        browser.sleep(5)

    def test_case_ticket(self, browser):
        # 加载Excel文件
        workbook = openpyxl.load_workbook('test_excel.xlsx')
        sheet = workbook.active
        # 遍历Excel文件中的测试用例
        for row in sheet.iter_rows(min_row=9, values_only=True):
            # 读取测试用例信息
            case_id = int(row[0])
            action = row[3]
            param1 = row[4]
            param2 = row[5]
            if case_id == 2:
                if action == 'name':
                    print(f"Test Case {case_id}: {param1}")
                elif action == 'goto':
                    browser.goto(param1)
                elif action == 'input':
                    browser.input(param1, param2)
                elif action == 'click':
                    browser.click(param1)
                elif action == 'hide_div':
                    browser.hide_div(param1, param2)
            else:
                continue
        browser.driver.get_screenshot_as_file("case_ticket.png")
        browser.sleep(5)
        # 运行测试用例
    if __name__ == "__main__":
        test_case_ticket()