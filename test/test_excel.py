import time
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By

def test_excel():
    # 加载Excel文件
    workbook = openpyxl.load_workbook('test_excel.xlsx')
    sheet = workbook.active

    # 读取浏览器类型并初始化WebDriver
    browser_type = sheet.cell(row=2, column=5).value
    if browser_type == 'chrome':
        driver = webdriver.Chrome()  # 确保已安装ChromeDriver并配置好环境变量
    else:
        raise ValueError(f"Unsupported browser type: {browser_type}")
    driver.maximize_window()

    # 遍历Excel文件中的测试用例
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 跳过表头行（第一行）
        if row[0] is None:  # 假设第一行的用例字段为空，用于判断表头
            continue

        # 读取测试用例信息
        case_id = int(row[0])
        action = row[3]
        param1 = row[4]
        param2 = row[5]

        # 根据关键字执行相应的操作
        if action == 'name':  # 设置用例名称（此处仅为记录，不执行实际操作）
            print(f"Test Case {case_id}: {param1}")
        elif action == 'goto':
            driver.get(param1)
        elif action == 'input':
            element = driver.find_element(By.XPATH, param1)
            element.clear()  # 清除输入框中的现有内容
            element.send_keys(param2)
        elif action == 'click':
            driver.find_element(By.XPATH, param1).click()
    driver.get_screenshot_as_file("a.png")
    time.sleep(5)
    # 测试完成后关闭浏览器
    driver.quit()

# 运行测试用例
if __name__ == "__main__":
    test_excel()
